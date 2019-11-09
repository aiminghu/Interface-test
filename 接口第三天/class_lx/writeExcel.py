#！ /usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2019/11/5 0005 下午 21:05 
# @Author : aiminhu
# @File : writeExcel.py 
# @Software: PyCharm
import openpyxl

# xlrd和xlwt处理的是xls文件，单个sheet最大行数是65535，如果有更大需要的，建议使用openpyxl函数，最大行数达到1048576。
# 如果数据量超过65535就会遇到：ValueError: row index was 65536, not allowed by .xls format

'''
#使用openpyxl库读写excel
# 1、打开excel
filename = r'E:\新建文件夹\zuoye\接口第三天\class_lx\\testCase.xlsx'
inwb = openpyxl.load_workbook(filename) #读文件
# 2、获取打开的excel的sheet内容
sheetnames = inwb.get_sheet_names() #获取读文件所有sheet，通过sheet页名字的方式
ws = inwb.get_sheet_by_name(sheetnames[0]) #获取第一个sheet页的内容
# 3、获取sheet的最大行数和列数
rows = ws.max_row
cols = ws.max_column
# 4、获取某个单元格的值
print(ws.cell(1,1).value)
# 5、打开将写的表并添加sheet
outwb = openpyxl.Workbook() #打开一个将写的文件
outws = outwb.create_sheet(index=0) #在将写的文件创建sheet
# 6、保存
saveExcel = "F:\\test\\files\\test.xlsx"
outwb.save(saveExcel) #一定要记得保存
'''
def readExcel():  #不太理解这里为啥要加self
    filename = r'E:\新建文件夹\zuoye\接口第三天\class_lx\\testCase.xlsx'
    inwb = openpyxl.load_workbook(filename) #读文件
    sheetnames = inwb.get_sheet_names()
    ws = inwb.get_sheet_by_name(sheetnames[0]) #获取第一个sheet内容
    # print(ws)

    #获取sheet的最大行数和列数
    rows = ws.max_row
    cols = ws.max_column
    for r in range(1,rows):
        for c in range(1,cols):
            pass
            # print(ws.cell(r,c).value)
        if r == 10:
            break


def writeExcel(): #不太理解这个层级关系，为什么writeExcel这个方法在里面
    outwb = openpyxl.Workbook() #打开一个将写的文件
    outws = outwb.create_sheet(index=0) #在将写的文件创建sheet
    for row in range(1,700):
        for col in range(1,7):
            outws.cell(row,col).value = row*2 #写文件
        print(row)
    saveExcel = "F:\\test\\files\\test.xlsx"
    outwb.save(saveExcel)

if __name__ == '__main__':
    readExcel()
    writeExcel()